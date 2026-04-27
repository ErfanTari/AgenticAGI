import os
import re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

base_output_dir = "outputs/benchmark-excel"
output_dir = base_output_dir
i = 0
while os.path.exists(output_dir):
    i += 1
    output_dir = f"{base_output_dir}-{i}"

os.makedirs(output_dir)

workbook = Workbook()

# Detailed Data Sheet
detailed_sheet = workbook.active
detailed_sheet.title = 'Detailed Data'

# Detailed Data Headers
headers = ['Product', 'Sales', 'Region', 'Date']
detailed_sheet.append(headers)

# Apply header formatting
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
for col_num, header_text in enumerate(headers, 1):
    cell = detailed_sheet.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    detailed_sheet.column_dimensions[get_column_letter(col_num)].width = 15

# Sample Detailed Data
sample_data = [
    ['Laptop', 1200, 'North', date(2023, 1, 15)],
    ['Mouse', 25, 'South', date(2023, 1, 16)],
    ['Keyboard', 75, 'East', date(2023, 1, 17)],
    ['Monitor', 300, 'West', date(2023, 1, 18)],
    ['Laptop', 1500, 'North', date(2023, 1, 19)],
    ['M
...[2149 chars elided — use file_reader or skill_schema to fetch if needed]...
_idx - 1, max_col=2)
categories = Reference(summary_sheet, min_col=1, min_row=2, max_row=summary_row_idx - 1)

chart.add_data(data, titles_from_data=False)
chart.set_categories(categories)

# Embed the chart in the 'Summary' sheet
summary_sheet.add_chart(chart, "A5") # Place chart at cell A5

# Save the workbook
file_path = os.path.join(output_dir, 'benchmark_report.xlsx')
workbook.save(file_path)